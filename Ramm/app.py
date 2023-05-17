from flask import Flask, render_template, request, send_file
import pandas as pd
import io
from openpyxl import Workbook


app = Flask(__name__)

def generate_excel(data, components):
    # Create a new Excel workbook
    wb = Workbook()

    # Add a sheet to the workbook and set its title
    ws = wb.active
    ws.title = "Input Details"

    # Add data to the sheet
    ws["A1"] = "Bundle_Number"
    ws["B1"] = data["Bundle_Number"]

    ws["A2"] = "Teacher"
    ws["B2"] = data["Teacher"]

    ws["A3"] = "Academic_year"
    ws["B3"] = data["Academic_year"]

    ws["A4"] = "Semester"
    ws["B4"] = data["Semester"]

    ws["A5"] = "Branch"
    ws["B5"] = data["Branch"]

    ws["A6"] = "Batch"
    ws["B6"] = data["Batch"]

    ws["A7"] = "Section"
    ws["B7"] = data["Section"]

    ws["A8"] = "Subject_Code"
    ws["B8"] = data["Subject_Code"]

    ws["A9"] = "Subject_Name"
    ws["B9"] = data["Subject_Name"]

    ws["A10"] = "Number_of_Students"
    ws["B10"] = data["Number_of_Students"]

    ws["A11"] = "Number_of_COs"
    ws["B11"] = data["Number_of_COs"]

    ws["A12"] = "Internal"
    ws["B12"] = data["Internal"]

    ws["A13"] = "External"
    ws["B13"] = data["External"]

    ws["A14"] = "Direct"
    ws["B14"] = data["Direct"]

    ws["A15"] = "Indirect"
    ws["B15"] = data["Indirect"]

    row = 16
    for component in components:
        ws[f"A{row}"] = component["name"]
        ws[f"B{row}"] = component["questions"]
        row += 1


    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        data = {
            "Bundle_Number": request.form["bundle_number"],
            "Teacher": request.form["teacher"],
            "Academic_year": request.form["academic_year"],
            "Semester": int(request.form["semester"]),
            "Branch": request.form["branch"],
            "Batch": int(request.form["batch"]),
            "Section": request.form["section"],
            "Subject_Code": request.form["subject_code"],
            "Subject_Name": request.form["subject_name"],
            "Number_of_Students": int(request.form["number_of_students"]),
            "Number_of_COs": int(request.form["number_of_cos"]),
            "Internal": int(request.form["internal"]),
            "External": int(request.form["external"]),
            "Direct": int(request.form["direct"]),
            "Indirect": int(request.form["indirect"]),
            "defaulThreshold": int(request.form["defaultThreshold"])
        }
        
        components = []
        num_components = int(request.form["number_of_components"])
        # for i in range(num_components):
        #     component_name = request.form[f"component_name_{i+1}"]
        #     num_questions = int(request.form[f"number_of_questions_{i+1}"])
        #     components.append({"name": component_name, "questions": num_questions})
        for i in range(num_components):
            component_name = request.form[f"component_{i}_name"]
            num_questions_str = request.form[f"component_{i}_num_questions"]
            num_questions = int(num_questions_str) if num_questions_str else 0  # changed this line
            components.append({"name": component_name, "questions": num_questions})



        excel_file = generate_excel(data, components)
        return send_file(
            excel_file,
            download_name="output.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)

